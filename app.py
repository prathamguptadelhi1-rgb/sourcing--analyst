import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Sourcing Controller AI", layout="wide")
st.title("🏭 Automated Sourcing Controller")
st.markdown("### Upload Purchase Register for 6-Lever Cost Analysis")

uploaded_file = st.file_uploader("Upload CSV/Excel", type=["csv", "xlsx"])

if uploaded_file:
    # ── 1. Load & clean ─────────────────────────────────────────────────────
    df = (
        pd.read_csv(uploaded_file, skiprows=3)
        if uploaded_file.name.endswith(".csv")
        else pd.read_excel(uploaded_file, skiprows=3)
    )

    for col in ["Value IN INR", "INR Price", "Quantity"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["Origin"] = df["Country"].apply(
        lambda x: "Domestic" if str(x).strip().lower() == "india" else "Import"
    )

    # ── 2. Lever calculations (dashboard) ───────────────────────────────────
    price_gap = df.groupby("Item No.").agg(
        Min_Price=("INR Price", "min"),
        Avg_Price=("INR Price", "mean"),
        Total_Qty=("Quantity", "sum"),
    )
    price_gap["Savings"] = (
        (price_gap["Avg_Price"] - price_gap["Min_Price"]) * price_gap["Total_Qty"]
    )

    import_val = df[df["Origin"] == "Import"]["Value IN INR"].sum()
    total_spend = df["Value IN INR"].sum()

    # ── 3. Dashboard ─────────────────────────────────────────────────────────
    st.header("📊 Consolidated Findings")
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Spend", f"₹{total_spend:,.0f}")
    c2.metric(
        "Savings Opportunity",
        f"₹{price_gap['Savings'].sum() + (import_val * 0.05):,.0f}",
    )
    c3.metric("Import Share", f"{(import_val / total_spend) * 100:.1f}%")

    st.subheader("💡 Recommendations")
    st.info(
        f"1. **Price Discovery:** Negotiate with vendors for top SKUs to save ₹{price_gap['Savings'].sum():,.0f}.\n"
        f"2. **Localization:** Move high-value imports to India to save approx. ₹{import_val * 0.05:,.0f}.\n"
        f"3. **SOB Allocation:** Shift 15% volume to the 'Min Price' vendors identified in the report."
    )

    # ── 4. Build 5 analysis DataFrames ──────────────────────────────────────

    # Sheet 1 – Category Spend
    sheet1 = (
        df.groupby("Category Description")["Value IN INR"]
        .sum()
        .reset_index()
        .sort_values("Value IN INR", ascending=False)
        .rename(columns={"Value IN INR": "Total Value (INR)"})
    )
    sheet1["% of Total Spend"] = sheet1["Total Value (INR)"] / total_spend * 100

    # Sheet 2 – Category vs Domestic / Import %
    s2_raw = (
        df.groupby(["Category Description", "Origin"])["Value IN INR"]
        .sum()
        .unstack(fill_value=0)
        .reset_index()
    )
    for col in ["Domestic", "Import"]:
        if col not in s2_raw.columns:
            s2_raw[col] = 0
    s2_raw["Total Value (INR)"] = s2_raw["Domestic"] + s2_raw["Import"]
    s2_raw["Domestic %"] = s2_raw["Domestic"] / s2_raw["Total Value (INR)"] * 100
    s2_raw["Import %"]   = s2_raw["Import"]   / s2_raw["Total Value (INR)"] * 100
    sheet2 = s2_raw[
        ["Category Description", "Domestic", "Import", "Total Value (INR)", "Domestic %", "Import %"]
    ].sort_values("Total Value (INR)", ascending=False)
    sheet2.columns = [
        "Category Description", "Domestic Value (INR)", "Import Value (INR)",
        "Total Value (INR)", "Domestic %", "Import %",
    ]

    # Sheet 3 – Category Avg Price & Total Value
    sheet3 = (
        df.groupby("Category Description")
        .agg(
            Avg_INR_Price=("INR Price", "mean"),
            Total_Value_INR=("Value IN INR", "sum"),
        )
        .reset_index()
        .sort_values("Total_Value_INR", ascending=False)
    )
    sheet3.columns = ["Category Description", "Avg INR Price (per unit)", "Total Value (INR)"]

    # Sheet 4 – Vendors per Category (descending by Qty)
    sheet4 = (
        df.groupby(["Category Description", "Vendor Name"])
        .agg(
            Total_Qty=("Quantity", "sum"),
            Total_Value_INR=("Value IN INR", "sum"),
            Avg_Price_per_kg=("INR Price", "mean"),
        )
        .reset_index()
        .sort_values(["Category Description", "Total_Qty"], ascending=[True, False])
    )
    sheet4.columns = [
        "Category Description", "Vendor Name",
        "Total Quantity (Kg)", "Total Value (INR)", "Avg Price per Kg (INR)",
    ]

    # Sheet 5 – RM & BP Category Summary
    sheet5 = (
        df[df["RM Category"].isin(["RM", "BP"])]
        .groupby(["RM Category", "Category Description"])
        .agg(
            Vendor_Count=("Vendor Name", "nunique"),
            Total_Value=("Value IN INR", "sum"),
        )
        .reset_index()
    )
    sheet5["Avg Purchase per Vendor (INR)"] = (
        sheet5["Total_Value"] / sheet5["Vendor_Count"]
    )
    sheet5.columns = [
        "Type (RM/BP)", "Category Description",
        "Vendor Count", "Total Value (INR)", "Avg Purchase per Vendor (INR)",
    ]
    sheet5 = sheet5.sort_values(["Type (RM/BP)", "Total Value (INR)"], ascending=[True, False])

    # ── 5. Write styled Excel workbook ──────────────────────────────────────

    def _hdr_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def _thin_border():
        s = Side(border_style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def _write_sheet(wb, title, dataframe, header_color="1F4E79", num_cols=None):
        ws = wb.create_sheet(title)
        num_cols = num_cols or []

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dataframe.columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
        title_cell.fill = _hdr_fill(header_color)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Header row
        for col_idx, col_name in enumerate(dataframe.columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = _hdr_fill("2E75B6")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _thin_border()
        ws.row_dimensions[2].height = 30

        # Data rows
        for row_idx, row_data in enumerate(dataframe.itertuples(index=False), start=3):
            fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Arial", size=10)
                cell.fill = _hdr_fill(fill_color)
                cell.border = _thin_border()
                cell.alignment = Alignment(horizontal="right" if col_idx in num_cols else "left")
                # Number formatting
                if col_idx in num_cols:
                    col_name = dataframe.columns[col_idx - 1]
                    if "%" in col_name:
                        cell.number_format = "0.00%"
                        if isinstance(value, (int, float)):
                            cell.value = value / 100
                    elif "Price" in col_name or "Value" in col_name or "Avg" in col_name:
                        cell.number_format = '₹#,##0.00'
                    elif "Qty" in col_name or "Quantity" in col_name or "Count" in col_name:
                        cell.number_format = "#,##0.00"

        # Auto-fit column widths
        for col_idx, col_name in enumerate(dataframe.columns, start=1):
            max_len = max(
                len(str(col_name)),
                *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(3, ws.max_row + 1)),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # Freeze panes below header
        ws.freeze_panes = "A3"

    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Sheet configs: (title, df, header_hex, numeric_col_indices)
    sheets_cfg = [
        ("1. Category Spend",          sheet1, "1F4E79", [2, 3]),
        ("2. Domestic vs Import",       sheet2, "375623", [2, 3, 4, 5, 6]),
        ("3. Avg Price & Total Value",  sheet3, "7030A0", [2, 3]),
        ("4. Vendor Analysis",          sheet4, "833C00", [3, 4, 5]),
        ("5. RM & BP Summary",         sheet5, "C00000", [3, 4, 5]),
    ]

    for title, df_s, color, num_cols in sheets_cfg:
        _write_sheet(wb, title, df_s, header_color=color, num_cols=num_cols)

    wb.save(output)

    st.download_button(
        "📥 Download Full Analysis Report",
        output.getvalue(),
        "Sourcing_Analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
